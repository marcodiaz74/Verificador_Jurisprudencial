# utils.py
# Funciones auxiliares: extracción de texto de documentos y exportación de reportes.

import io
import re
import pandas as pd
from extractor import Cita, ETIQUETAS_TIPO


# ──────────────────────────────────────────────────────────────────────────────
# EXTRACCIÓN DE TEXTO SEGÚN FORMATO
# ──────────────────────────────────────────────────────────────────────────────

def leer_pdf(archivo_bytes: bytes) -> str:
    """Extrae texto de un archivo PDF usando pdfplumber."""
    import pdfplumber
    texto = []
    with pdfplumber.open(io.BytesIO(archivo_bytes)) as pdf:
        for pagina in pdf.pages:
            contenido = pagina.extract_text()
            if contenido:
                texto.append(contenido)
    return "\n".join(texto)


def leer_docx(archivo_bytes: bytes) -> str:
    """Extrae texto de un archivo DOCX usando python-docx."""
    from docx import Document
    doc = Document(io.BytesIO(archivo_bytes))
    parrafos = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(parrafos)


def leer_txt(archivo_bytes: bytes) -> str:
    """Lee texto plano intentando distintas codificaciones."""
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return archivo_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return archivo_bytes.decode("utf-8", errors="replace")


def extraer_texto(archivo_bytes: bytes, nombre_archivo: str) -> str:
    """Dispatcher: selecciona el lector según la extensión del archivo."""
    extension = nombre_archivo.rsplit(".", 1)[-1].lower()
    if extension == "pdf":
        return leer_pdf(archivo_bytes)
    elif extension in ("docx", "doc"):
        return leer_docx(archivo_bytes)
    elif extension == "txt":
        return leer_txt(archivo_bytes)
    else:
        raise ValueError(f"Formato no soportado: .{extension}")


def limpiar_texto(texto: str) -> str:
    """Normaliza espacios y caracteres especiales sin alterar el contenido jurídico."""
    # Normalizar guiones tipográficos a guión estándar
    texto = texto.replace("\u2013", "-").replace("\u2014", "-").replace("\u2010", "-")
    # Normalizar espacios múltiples
    texto = re.sub(r" {2,}", " ", texto)
    # Eliminar líneas completamente vacías consecutivas
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    return texto.strip()


# ──────────────────────────────────────────────────────────────────────────────
# CONVERSIÓN A DATAFRAME
# ──────────────────────────────────────────────────────────────────────────────

ETIQUETAS_ESTADO = {
    "encontrada": "✅ Encontrada",
    "no_encontrada": "❌ No encontrada",
    "no_verificable": "⚠️ No verificable",
    "pendiente": "🔄 Pendiente",
}

ETIQUETAS_TIPO_CORTO = {
    "norma": "Norma",
    "corte_constitucional": "Corte Constitucional",
    "corte_suprema": "Corte Suprema",
    "consejo_estado": "Consejo de Estado",
    "doctrina": "Doctrina",
}


def citas_a_dataframe(citas: list[Cita]) -> pd.DataFrame:
    """Convierte la lista de citas a un DataFrame de pandas para mostrar en Streamlit."""
    filas = []
    for c in citas:
        filas.append({
            "Tipo": ETIQUETAS_TIPO_CORTO.get(c.tipo, c.tipo),
            "Subtipo": c.subtipo,
            "Cita en el documento": c.texto_original,
            "Referencia normalizada": c.referencia,
            "Estado": ETIQUETAS_ESTADO.get(c.estado, c.estado),
            "Fuente consultada": c.fuente,
            "Enlace": c.enlace,
        })
    return pd.DataFrame(filas)


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN A EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def exportar_excel(citas: list[Cita], nombre_documento: str = "") -> bytes:
    """
    Genera un archivo Excel (.xlsx) con el reporte completo de verificación.
    Retorna los bytes del archivo listo para descargar.
    """
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Hoja 1: Reporte completo ──
    ws = wb.active
    ws.title = "Reporte Citas"

    # Colores por estado
    COLOR_VERDE = "C6EFCE"
    COLOR_ROJO = "FFC7CE"
    COLOR_AMARILLO = "FFEB9C"
    COLOR_GRIS = "F2F2F2"
    COLOR_AZUL_OSCURO = "1F3864"

    # Encabezado del documento
    ws.merge_cells("A1:G1")
    celda_titulo = ws["A1"]
    celda_titulo.value = "VERIFICADOR DE CITAS JURÍDICAS COLOMBIANAS"
    celda_titulo.font = Font(bold=True, color="FFFFFF", size=13)
    celda_titulo.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if nombre_documento:
        ws.merge_cells("A2:G2")
        ws["A2"].value = f"Documento analizado: {nombre_documento}"
        ws["A2"].font = Font(italic=True, size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

    # Aviso Habeas Data
    ws.merge_cells("A3:G3")
    ws["A3"].value = (
        "⚖️ Datos personales anonimizados conforme a la Ley 1581/2012 (Habeas Data). "
        "Este reporte es orientativo y no reemplaza el criterio jurídico profesional."
    )
    ws["A3"].font = Font(italic=True, size=9, color="7F7F7F")
    ws["A3"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[3].height = 20

    # Encabezados de columna
    columnas = [
        "Tipo", "Subtipo", "Cita en el documento",
        "Referencia normalizada", "Estado", "Fuente consultada", "Enlace"
    ]
    anchos = [18, 20, 40, 30, 20, 35, 45]
    fila_encabezado = 5

    borde_fino = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, (col_nombre, ancho) in enumerate(zip(columnas, anchos), start=1):
        celda = ws.cell(row=fila_encabezado, column=col_idx, value=col_nombre)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="2E4A7A")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde_fino
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[fila_encabezado].height = 22

    # Filas de datos
    for fila_idx, cita in enumerate(citas, start=fila_encabezado + 1):
        datos = [
            ETIQUETAS_TIPO_CORTO.get(cita.tipo, cita.tipo),
            cita.subtipo,
            cita.texto_original,
            cita.referencia,
            ETIQUETAS_ESTADO.get(cita.estado, cita.estado),
            cita.fuente,
            cita.enlace,
        ]

        # Color de fondo según estado
        if cita.estado == "encontrada":
            color_fondo = COLOR_VERDE
        elif cita.estado == "no_encontrada":
            color_fondo = COLOR_ROJO
        elif cita.estado == "no_verificable":
            color_fondo = COLOR_AMARILLO
        else:
            color_fondo = COLOR_GRIS

        for col_idx, valor in enumerate(datos, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.fill = PatternFill("solid", fgColor=color_fondo)
            celda.alignment = Alignment(wrap_text=True, vertical="center")
            celda.border = borde_fino
            # Enlace como hipervínculo
            if col_idx == 7 and valor:
                celda.hyperlink = valor
                celda.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = f"A{fila_encabezado + 1}"

    # ── Hoja 2: Estadísticas ──
    ws2 = wb.create_sheet("Estadísticas")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    total = len(citas)
    encontradas = sum(1 for c in citas if c.estado == "encontrada")
    no_encontradas = sum(1 for c in citas if c.estado == "no_encontrada")
    no_verificables = sum(1 for c in citas if c.estado == "no_verificable")

    ws2["A1"].value = "ESTADÍSTICAS DE VERIFICACIÓN"
    ws2["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    ws2.merge_cells("A1:B1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    estadisticas = [
        ("Total de citas detectadas", total),
        ("✅ Encontradas", encontradas),
        ("❌ No encontradas", no_encontradas),
        ("⚠️ No verificables", no_verificables),
        ("% Verificadas", f"{(encontradas + no_encontradas) / total * 100:.1f}%" if total else "0%"),
        ("% Válidas", f"{encontradas / total * 100:.1f}%" if total else "0%"),
        ("% Con error/no verificable", f"{no_verificables / total * 100:.1f}%" if total else "0%"),
    ]

    por_tipo = {}
    for cita in citas:
        etiqueta = ETIQUETAS_TIPO_CORTO.get(cita.tipo, cita.tipo)
        por_tipo[etiqueta] = por_tipo.get(etiqueta, 0) + 1

    for fila_idx, (etiqueta, valor) in enumerate(estadisticas, start=2):
        ws2.cell(row=fila_idx, column=1, value=etiqueta).alignment = Alignment(wrap_text=True)
        ws2.cell(row=fila_idx, column=2, value=valor).alignment = Alignment(horizontal="center")

    fila_tipo_inicio = len(estadisticas) + 3
    ws2.cell(row=fila_tipo_inicio, column=1, value="Distribución por tipo").font = Font(bold=True)
    for i, (tipo, cantidad) in enumerate(por_tipo.items(), start=fila_tipo_inicio + 1):
        ws2.cell(row=i, column=1, value=tipo)
        ws2.cell(row=i, column=2, value=cantidad)

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# ESTADÍSTICAS
# ──────────────────────────────────────────────────────────────────────────────

def calcular_estadisticas(citas: list[Cita]) -> dict:
    total = len(citas)
    if total == 0:
        return {"total": 0, "encontradas": 0, "no_encontradas": 0, "no_verificables": 0,
                "pct_verificadas": 0, "pct_validas": 0, "pct_error": 0, "por_tipo": {}}

    encontradas = sum(1 for c in citas if c.estado == "encontrada")
    no_encontradas = sum(1 for c in citas if c.estado == "no_encontrada")
    no_verificables = sum(1 for c in citas if c.estado == "no_verificable")

    por_tipo: dict[str, dict] = {}
    for cita in citas:
        tipo = ETIQUETAS_TIPO_CORTO.get(cita.tipo, cita.tipo)
        if tipo not in por_tipo:
            por_tipo[tipo] = {"total": 0, "encontradas": 0, "no_encontradas": 0, "no_verificables": 0}
        por_tipo[tipo]["total"] += 1
        por_tipo[tipo][cita.estado.replace("-", "_")] = por_tipo[tipo].get(cita.estado, 0) + 1

    return {
        "total": total,
        "encontradas": encontradas,
        "no_encontradas": no_encontradas,
        "no_verificables": no_verificables,
        "pct_verificadas": round((encontradas + no_encontradas) / total * 100, 1),
        "pct_validas": round(encontradas / total * 100, 1),
        "pct_error": round(no_verificables / total * 100, 1),
        "por_tipo": por_tipo,
    }
